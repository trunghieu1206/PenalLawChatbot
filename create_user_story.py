import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

target_file = "VNPLaw_UserStory.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "User Story"

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="00A2E8", end_color="00A2E8", fill_type="solid")  # Blue header
actor_fill_user = PatternFill(start_color="B5E6A1", end_color="B5E6A1", fill_type="solid") # Light green
actor_fill_admin = PatternFill(start_color="4DB8FF", end_color="4DB8FF", fill_type="solid") # Light blue
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
top_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Headers
headers = ["Actor", "User story", "User story description", "Task", "Estimate", "Task description", "Estimate "]
for col_idx, header in enumerate(headers, start=2):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 15 # Actor
ws.column_dimensions['C'].width = 25 # User story
ws.column_dimensions['D'].width = 40 # User story description
ws.column_dimensions['E'].width = 30 # Task
ws.column_dimensions['F'].width = 10 # Estimate
ws.column_dimensions['G'].width = 40 # Task description
ws.column_dimensions['H'].width = 10 # Estimate

# Data Definition
# Format: (Actor, [ (User Story, US Desc, [ (Task, Task Desc) ]) ])
data = [
    ("User / Guest", [
        (
            "Phân tích vụ án hình sự",
            "Người dùng nhập thông tin vụ án (hành vi, ngày phạm tội...) và nhận kết quả phân tích theo vai trò (Thẩm phán, LS bào chữa, LS bị hại).",
            [
                ("Tạo giao diện chat", "Giao diện chat cho người dùng nhập văn bản (case content) và chọn vai trò"),
                ("Phân loại ý định (classify_intent)", "Sử dụng LLM phân loại câu hỏi (new_case, followup, casual)"),
                ("Trích xuất thông tin (extract_facts)", "Tạo Node trích xuất hành vi, ngày phạm tội, tình tiết tăng nặng/giảm nhẹ thành dạng JSON"),
                ("Kiểm tra thông tin (clarification_check)", "Kiểm tra các trường bắt buộc (hành vi, ngày phạm tội > 2000)"),
                ("Tạo truy vấn tìm kiếm", "Tạo 3 truy vấn semantic (behavior, circumstance, evidence) từ dữ liệu đã trích xuất"),
                ("Tìm kiếm semantic + pinned", "Truy xuất Milvus Lite (semantic) kết hợp query cứng (pinned_fetch) các điều khoản thủ tục"),
                ("Xếp hạng kết quả (rerank)", "Sử dụng BGE-M3 Reranker để lấy top 5 tài liệu liên quan nhất"),
                ("Ánh xạ luật (map_laws)", "Sử dụng nguyên tắc thời hiệu (Điều 7 BLHS) để ánh xạ hành vi với điều luật tương ứng"),
                ("Sinh câu trả lời (generate)", "Tạo lập luận pháp lý theo góc nhìn của vai trò đã chọn (bias mode)")
            ]
        ),
        (
            "Tra cứu chi tiết điều luật",
            "Người dùng xem chi tiết nội dung văn bản luật trực tiếp từ các điều khoản được AI trích dẫn trong câu trả lời.",
            [
                ("Tạo API tra cứu luật", "API GET /api/laws/{articleNumber} tra cứu từ PostgreSQL"),
                ("Xử lý phiên bản luật", "Hệ thống tự động hiển thị phiên bản luật áp dụng tại thời điểm phạm tội (crimeDate) ở tab mặc định"),
                ("Giao diện Sidebar", "Thiết kế LawSidebar hiển thị nội dung điều luật dạng tab (các phiên bản lịch sử)")
            ]
        ),
        (
            "Luyện tập phân tích pháp lý",
            "Người dùng đóng vai chuyên gia pháp lý, tự viết phân tích và gửi cho AI chấm điểm, nhận xét.",
            [
                ("Giao diện Practice Mode", "Màn hình TrainingPage có form nhập phân tích của người dùng"),
                ("API chấm điểm", "API POST /practice/evaluate gọi AI service"),
                ("Node chấm điểm (rebuttal_node)", "Sử dụng LLM so sánh phân tích của người dùng với kết quả chuẩn của hệ thống (chấm điểm 0-100, ưu điểm, nhược điểm)")
            ]
        ),
        (
            "Phản hồi kết quả AI",
            "Người dùng đánh giá chất lượng câu trả lời của AI (Đúng/Sai) để cải thiện hệ thống.",
            [
                ("Tạo nút đánh giá UI", "Thêm nút Thích / Không thích vào từng MessageBubble"),
                ("Tạo form gửi phản hồi", "Hiển thị popup cho phép người dùng điền bình luận giải thích"),
                ("API gửi phản hồi", "API POST /api/admin/feedback lưu đánh giá vào database")
            ]
        ),
        (
            "Quản lý tài khoản và Lịch sử",
            "Người dùng đăng ký/đăng nhập để lưu trữ vĩnh viễn các phiên trò chuyện, có thể xem lại trên thiết bị khác.",
            [
                ("Tạo luồng Đăng ký/Đăng nhập", "Giao diện LoginPage, RegisterPage và xử lý JWT Token"),
                ("Lưu trữ phiên trò chuyện", "Lưu trữ cấu trúc ChatSession và ChatMessage vào PostgreSQL"),
                ("Giao diện Sidebar lịch sử", "Hiển thị danh sách các phiên trò chuyện cũ bên trái màn hình")
            ]
        )
    ]),
    ("Admin", [
        (
            "Theo dõi thống kê hệ thống",
            "Quản trị viên xem số lượng người dùng truy cập, tổng số vụ án đã tư vấn, và phân loại theo vai trò trên Dashboard.",
            [
                ("Thiết kế Dashboard UI", "Màn hình AdminPage hiển thị biểu đồ và các thẻ số liệu thống kê tổng quan"),
                ("Theo dõi Visitor (Track Visit)", "Tạo thuật toán track unique daily visitor (localStorage UUID + deduplication ở DB)"),
                ("API lấy dữ liệu thống kê", "API GET /api/admin/stats và /api/home trả về số liệu tổng hợp")
            ]
        ),
        (
            "Quản lý phản hồi người dùng",
            "Quản trị viên duyệt các phản hồi Đúng/Sai, xem chi tiết lịch sử đoạn chat để phân tích lỗi của AI.",
            [
                ("Giao diện danh sách phản hồi", "Bảng hiển thị các phản hồi của người dùng với trạng thái (Cần xem xét / Đã xem xét)"),
                ("Chức năng xem Context", "Tính năng hiển thị toàn bộ lịch sử tin nhắn của phiên trò chuyện bị báo lỗi"),
                ("API cập nhật trạng thái", "API PATCH /api/admin/feedback/{id}/status để chuyển trạng thái xử lý")
            ]
        )
    ])
]

current_row = 3
for actor, stories in data:
    start_actor_row = current_row
    
    for story, story_desc, tasks in stories:
        start_story_row = current_row
        
        for task, task_desc in tasks:
            ws.cell(row=current_row, column=5, value=task).alignment = top_align
            ws.cell(row=current_row, column=7, value=task_desc).alignment = top_align
            
            for c in range(2, 9):
                ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1
            
        # Merge Story columns
        ws.merge_cells(start_row=start_story_row, start_column=3, end_row=current_row-1, end_column=3)
        ws.merge_cells(start_row=start_story_row, start_column=4, end_row=current_row-1, end_column=4)
        
        cell_story = ws.cell(row=start_story_row, column=3, value=story)
        cell_story.alignment = top_align
        cell_story.border = thin_border
        
        cell_desc = ws.cell(row=start_story_row, column=4, value=story_desc)
        cell_desc.alignment = top_align
        cell_desc.border = thin_border

    # Merge Actor column
    ws.merge_cells(start_row=start_actor_row, start_column=2, end_row=current_row-1, end_column=2)
    cell_actor = ws.cell(row=start_actor_row, column=2, value=actor)
    cell_actor.alignment = center_align
    cell_actor.border = thin_border
    
    # Set Actor color
    fill_color = actor_fill_user if actor == "User / Guest" else actor_fill_admin
    for r in range(start_actor_row, current_row):
        ws.cell(row=r, column=2).fill = fill_color

# Apply borders to empty estimate columns
for r in range(3, current_row):
    ws.cell(row=r, column=6).border = thin_border
    ws.cell(row=r, column=8).border = thin_border

wb.save(target_file)
print(f"Successfully created {target_file}")
